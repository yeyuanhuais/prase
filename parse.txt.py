import os
import re
import openpyxl
from openpyxl.styles import PatternFill
import json

# 包长度
PACKAGE_LENGTH = 20
# 执行次数
NUMBER_OF_EXECUTIONS = 100
# 输出提示
print("开始解析")

# 文件路径
LOG_FILE_PATH = "./50/info.24-12-26.0.log"
LOG_FILE_BACK_PATH = "./50/backmanage_log.txt"

# 创建Excel工作簿
workbook = openpyxl.Workbook()
sheet = workbook.active
# 设置背景颜色
highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 设置为黄色背景

# 初始化字典
# value_mapping {plc地址：值}
# value_last_mapping {plc地址：值} 包的最后一个值
# ext_mapping {ext：plc地址}
# addr_mapping {plc地址：触发次数}
# id_mapping {内部变量地址：触发次数}
# plc_mapping {plc地址：内部变量地址数}
addr_mapping, ext_mapping, value_mapping, id_mapping, plc_mapping,value_last_mapping, = {}, {}, {}, {},{},{}

initial_null = ["" for _ in range(NUMBER_OF_EXECUTIONS+1)]
# 初始的 addr 和 id 列表
initial_addrs = [2000, 4000, 6000, 8000, 0]
initial_ids = [27499, 29499, 31499, 33499, 4039]
initial_ids_addrs = [27499, 29499, 31499, 33499, 4039]
# 更新字典映射
def update_mappings():
    for i in range(NUMBER_OF_EXECUTIONS):
        for index, addr in enumerate(initial_addrs):
            addr_mapping[addr] = i + 1
            plc_mapping[addr] = initial_ids_addrs[index]
            initial_addrs[index] += PACKAGE_LENGTH
            initial_ids_addrs[index] += PACKAGE_LENGTH
        for index, addr in enumerate(initial_ids):
            id_mapping[addr] = i + 1
            initial_ids[index] += PACKAGE_LENGTH
update_mappings()

# 定义数据模板
parsed_data = [
    [""],
    ["采集端采集到触发变量改变 16059"],
    ["采集端发送触发变量改变 16059"],
    ["www接收到改变 16059"],
    ["node接收到改变 16059"],
]
for item in parsed_data:
    item.extend(initial_null)
# 其他数据模板
data_templates = {
    "front_end_script_sending": [["前端脚本发送0x30 y1", ""], ["前端脚本发送0x30 y2", ""], ["前端脚本发送0x30 y3", ""], ["前端脚本发送0x30 y4", ""], ["前端脚本发送0x30 x", ""]],
    "node_script_sending": [["node脚本发送0x30 y1", ""], ["node脚本发送0x30 y2", ""], ["node脚本发送0x30 y3", ""], ["node脚本发送0x30 y4", ""], ["node脚本发送0x30 x", ""]],
    "received_the_collection_end": [["采集端收到0x30 y1", ""], ["采集端收到0x30 y2", ""], ["采集端收到0x30 y3", ""], ["采集端收到0x30 y4", ""], ["采集端收到0x30 x", ""]],
    "collect_plc": [["采集端0x30开始读plc数据时间值 y1", ""], ["采集端0x30开始读plc数据时间值 y2", ""], ["采集端0x30开始读plc数据时间值 y3", ""], ["采集端0x30开始读plc数据时间值 y4", ""], ["采集端0x30开始读plc数据时间值 x", ""]],
    "read_plc": [["采集端0x30读到plc y1", ""], ["采集端0x30读到plc y2", ""], ["采集端0x30读到plc y3", ""], ["采集端0x30读到plc y4", ""], ["采集端0x30读到plc x", ""]],
    "collection_end_sending": [["采集端发送0x30 y1", ""], ["采集端发送0x30 y2", ""], ["采集端发送0x30 y3", ""], ["采集端发送0x30 y4", ""], ["采集端发送0x30 x", ""]],
    "www_receive_data": [["www接收0x30 y1", ""], ["www接收0x30 y2", ""], ["www接收0x30 y3", ""], ["www接收0x30 y4", ""], ["www接收0x30 x", ""]],
    "node_script_receiving": [["node脚本接收0x30 y1", ""], ["node脚本接收0x30 y2", ""], ["node脚本接收0x30 y3", ""], ["node脚本接收0x30 y4", ""], ["node脚本接收0x30 x", ""]],
    "front_end_script_sending_0x0d": [["前端脚本发送0x0d y1", ""], ["前端脚本发送0x0d y2", ""], ["前端脚本发送0x0d y3", ""], ["前端脚本发送0x0d y4", ""], ["前端脚本发送0x0d x", ""]],
    "node_sending_0x0d": [["node脚本发送0x0d y1", ""], ["node脚本发送0x0d y2", ""], ["node脚本发送0x0d y3", ""], ["node脚本发送0x0d y4", ""], ["node脚本发送0x0d x", ""]],
    "received_the_collection_end_0x0d": [["采集端收到0x0d y1", ""], ["采集端收到0x0d y2", ""], ["采集端收到0x0d y3", ""], ["采集端收到0x0d y4", ""], ["采集端收到0x0d x", ""]],
    "collection_end_sending_0x01": [["采集端发送0x01 y1", ""], ["采集端发送0x01 y2", ""], ["采集端发送0x01 y3", ""], ["采集端发送0x01 y4", ""], ["采集端发送0x01 x", ""]],
    "www_receive_0x01": [["www收到0x01 y1", ""], ["www收到0x01 y2", ""], ["www收到0x01 y3", ""], ["www收到0x01 y4", ""], ["www收到0x01 x", ""]],
    "node_receive_0x01": [["node收到0x01 y1", ""], ["node收到0x01 y2", ""], ["node收到0x01 y3", ""], ["node收到0x01 y4", ""], ["node收到0x01 x", ""]],
    "front_end_receive_35537": [["前端收到35537"], ["前端渲染35537"]]
}

for key in data_templates:
    for item in data_templates[key]:
        item.extend(initial_null)

def find_key_by_value(my_dict, target_value):
    """根据value查找对应的key"""
    for key, value in my_dict.items():
        if value == target_value:
            return key
    return None  # 如果没有找到对应的值，返回None
# 根据 addr 的值映射到相应的 parsed_data 列表位置
def map_addr_to_index(addr):
    if 2000 <= addr < 4000:
        return 0
    elif 4000 <= addr < 6000:
        return 1
    elif 6000 <= addr < 8000:
        return 2
    elif 8000 <= addr < 10000:
        return 3
    else:
        return 4

# 根据 id 的值映射到相应的 parsed_data 列表位置
def map_id_to_index(id):
    if 27499 <= id < 29498:
        return 0
    elif 29499 <= id < 31498:
        return 1
    elif 31499 <= id < 33498:
        return 2
    elif 33499 <= id < 35498:
        return 3
    elif 4039 <= id < 6038:
        return 4

print("初始化完成")
# 处理日志文件
def process_log_file():
    if os.path.exists(LOG_FILE_PATH):
        with open(LOG_FILE_PATH, "r", encoding="utf-8") as file:
            for line in file:
                if "=============== on 0x01，" in line:
                    process_on_0x01(line)
                elif "=============== 脚本" in line and "发送" in line:
                    process_script_send(line)
                elif "===============send bm ，" in line:
                    process_send_bm(line)
                elif "===============on bm，" in line:
                    process_on_bm(line)
                elif "=============== 脚本" in line and "收到" in line:
                    process_script_receive(line)
                elif "=============== send BatchSetByFullNames，" in line:
                    process_send_batch_set_by_full_names(line)
                elif "=============== tcp receiveData，" in line:
                    process_tcp_receive_data(line)
                elif "=============== useVarTrigger 35537 ，" in line:
                    process_use_var_trigger(line)
                elif "=============== render 80ms ，" in line:
                    process_render(line)
    else:
        print("文件不存在！")
    if os.path.exists(LOG_FILE_BACK_PATH):
        with open(LOG_FILE_BACK_PATH, "r", encoding="utf-8",errors='replace') as file:
            for line in file:
                if "【读到plc触发变量的时间值】" in line:
                    process_on_plc(line,"back")
                elif "【0x01上报的时间值】" in line:
                    process_send_0x01(line)
                elif "【采集端收到0x30的时间值】" in line:
                    process_receive_0x30(line)
                elif "【0x30开始读plc数据时间值】" in line:
                    process_start_plc_0x30(line)
                elif "【0x30读到plc数据时间值】" in line:
                    process_read_plc_0x30(line)
                elif "【0x30上报的时间值】" in line:
                    process_send_0x30(line)
                elif "【采集端收到0x0d的时间值】" in line:
                    process_receive_0x0d(line)
    else:
        print("采集端日志文件不存在！")
    
    merge_data_templates()
    save_to_excel()
def extract_log_info(log_line ,text):
    # 定义正则表达式
    time_pattern = rf"【{re.escape(text)}】\[(\d+)\]"
    data_pattern = r"【Data】\[(.+?)\]"

    # 提取时间值
    time_match = re.search(time_pattern, log_line)
    time_value = time_match.group(1) if time_match else None

    # 提取Data值
    data_match = re.search(data_pattern, log_line)
    if data_match:
        data_match=data_match
    else:
        data_match = re.search(r'【Data】\[(.*)', log_line)
    data_value = data_match.group(1) if data_match else None

    # 构造成map对象
    result = {}
    if time_value:
        result["time"] = time_value
    if data_value:
        result["data"] = data_value
    
    return result

# 处理包含 "=============== on 0x01，" 0x01上报的时间值 的行
def process_send_0x01(line):
    data = extract_log_info(line,"0x01上报的时间值")
    if len(data):
        if "16059" in data["data"]:
            value_after_16059 = re.search(r'16059\|@(\d+)', data["data"]).group(1)
            if value_after_16059 in parsed_data[0]:
                parsed_data[2][parsed_data[0].index(value_after_16059)]=data["time"]
        process_message_0x01(data["data"], data["time"])    
# 处理消息内容并分割
def process_message_0x01(msg,time):
    for key in id_mapping:
        plc=find_key_by_value(plc_mapping,key)
        if len(value_mapping)>0 and str(plc) in value_mapping:
            value=value_mapping[str(plc)]
            if (str(key) + "|@"+str(value)) in msg:
                indextwo = id_mapping[int(key)]
                indexone = map_id_to_index(key)
                parsed_data_index = parsed_data[0].index(str(indextwo))
                data_templates["collection_end_sending_0x01"][indexone][parsed_data_index]=time



# 处理包含 读到plc触发变量的时间值 的行
def process_on_plc(line,back):
    if back:
        data = extract_log_info(line,"读到plc触发变量的时间值")
        if len(data):
            if "16059" in data["data"]:
                value_after_16059 = re.search(r'16059\|@(\d+)', data["data"]).group(1)
                # parsed_data[0][int(value_after_16059)+1]=(value_after_16059)
                parsed_data[1][parsed_data[0].index(value_after_16059)]=data["time"]
# 处理包含 采集端收到0x30的时间值 的行
def process_receive_0x30(line):
    data = extract_log_info(line,"采集端收到0x30的时间值")
    if len(data):
        if "addr" in data["data"]:
            value_after_addr = re.search(r'"addr"\s*:\s*(\d+)', data["data"]).group(1)
            indextwo = addr_mapping[int(value_after_addr)]
            indexone = map_addr_to_index(int(value_after_addr))
            if str(indextwo) in parsed_data[0]:
                data_templates["received_the_collection_end"][indexone][parsed_data[0].index(str(indextwo))]=data["time"]
# 处理包含 0x30开始读plc数据时间值 的行
def process_start_plc_0x30(line):
    data = extract_log_info(line,"0x30开始读plc数据时间值")
    if len(data):
        val = data["data"].split("|@")
        val=val[len(val)-2]
        value_after_addr=find_key_by_value(value_last_mapping,int(val))
        value_after_addr=int(value_after_addr)-PACKAGE_LENGTH+1
        indextwo = addr_mapping[value_after_addr]
        indexone = map_addr_to_index(value_after_addr)
        if str(indextwo) in parsed_data[0]:
            data_templates["collect_plc"][indexone][parsed_data[0].index(str(indextwo))]=data["time"]
# 处理包含 0x30读到plc数据时间值 的行
def process_read_plc_0x30(line):
    data = extract_log_info(line,"0x30读到plc数据时间值")
    if len(data):
        val = data["data"].split("|@")
        val=val[len(val)-2]
        value_after_addr=find_key_by_value(value_last_mapping,int(val))
        value_after_addr=int(value_after_addr)-PACKAGE_LENGTH+1
        indextwo = addr_mapping[value_after_addr]
        indexone = map_addr_to_index(value_after_addr)
        if str(indextwo) in parsed_data[0]:
            data_templates["read_plc"][indexone][parsed_data[0].index(str(indextwo))]=data["time"]

# 处理包含 0x30上报的时间值 的行
def process_send_0x30(line):
    data = extract_log_info(line,"0x30上报的时间值")
    if len(data):
        if "ext" in data["data"]:
            value_after_ext = re.search(r'"ext"\s*:"\s*(\d+)"', data["data"]).group(1)
            indexadd = ext_mapping[value_after_ext]
            indextwo = addr_mapping[int(indexadd)]
            indexone = map_addr_to_index(int(indexadd))
            if str(indextwo) in parsed_data[0]:
                data_templates["collection_end_sending"][indexone][parsed_data[0].index(str(indextwo))]=data["time"]
# 处理包含 采集端收到0x0d的时间值 的行
def process_receive_0x0d(line):
    data = extract_log_info(line,"采集端收到0x0d的时间值")
    if len(data):
        json_str = re.search(r'({.*})', data["data"]).group(1)
        onsend = json.loads(json_str)["data"]
        id = onsend.split("|@", 1)[0]
        matches = re.findall(r"(\d+)\|@(\d+)", onsend)
        last_key, last_value = matches[-1]
        if(last_key=="35537"):
            last_key, last_value = matches[-2]
        plc_key=find_key_by_value(plc_mapping,int(last_key)-PACKAGE_LENGTH+1)+PACKAGE_LENGTH-1
        if str(plc_key) in value_last_mapping and value_last_mapping[str(plc_key)]==int(last_value):
            indextwo = id_mapping[int(id)]
            indexone = map_id_to_index(int(id))
            if str(indextwo) in parsed_data[0]:
                data_templates["received_the_collection_end_0x0d"][indexone][parsed_data[0].index(str(indextwo))]=data["time"]
# 处理包含 "=============== on 0x01，" 0x01上报的时间值 的行
def process_on_0x01(line):
    msg_pattern = re.compile(r'msg: "([^"]+)"')
    msg_match = msg_pattern.search(line)
    if msg_match:
        msg = msg_match.group(1)
        if "16059" in msg:
            value_after_16059 = re.search(r'16059\|@(\d+)', msg).group(1)
            time = re.search(r'time:(\d+)', line).group(1)
            if(value_after_16059=="0"):
                parsed_data[4][1]=time
            else:
                print(value_after_16059,parsed_data[0])
                parsed_data[4][parsed_data[0].index(value_after_16059)]=time
        process_message_split(msg, line)    
# 处理消息内容并分割
def process_message_split(msg, line):
    for key in id_mapping:
        plc=find_key_by_value(plc_mapping,key)
        if len(value_mapping)>0 and str(plc) in value_mapping:
            value=value_mapping[str(plc)]
            if (str(key) + "|@"+str(value)) in msg:
                indextwo = id_mapping[int(key)]
                time = re.search(r'time:(\d+)', line).group(1)
                indexone = map_id_to_index(key)
                parsed_data_index = parsed_data[0].index(str(indextwo))
                data_templates["node_receive_0x01"][indexone][parsed_data_index]= time

# 处理包含 "=============== 脚本" 和 "发送" 的行
def process_script_send(line):
    match_addr = re.search(r'"addr"\s*:\s*(\d+)', line)
    if match_addr:
        value_after_addr = match_addr.group(1)
        if int(value_after_addr) in addr_mapping:
            indextwo = addr_mapping[int(value_after_addr)]
            time = re.search(r'time:(\d+)', line).group(1)
            if str(indextwo) in parsed_data[0]:
                if "y1" in line:
                    data_templates["front_end_script_sending"][0][parsed_data[0].index(str(indextwo))]=time
                elif "y2" in line:
                    data_templates["front_end_script_sending"][1][parsed_data[0].index(str(indextwo))]=time
                elif "y3" in line:
                    data_templates["front_end_script_sending"][2][parsed_data[0].index(str(indextwo))]=time
                elif "y4" in line:
                    data_templates["front_end_script_sending"][3][parsed_data[0].index(str(indextwo))]=time
                elif "x" in line:
                    data_templates["front_end_script_sending"][4][parsed_data[0].index(str(indextwo))]=time

# 处理包含 "===============send bm ，" 的行
def process_send_bm(line):
    parts = line.split('msg: ', 1)
    if len(parts) > 1:
        msg = parts[1]
        if "addr" in msg:
            value_after_addr = re.search(r'"addr"\s*:\s*(\d+)', msg).group(1)
            ext_value = re.search(r'"ext":"(\d+)"', line).group(1)
            ext_mapping[ext_value] = value_after_addr
            indextwo = addr_mapping[int(value_after_addr)]
            time = re.search(r'time:(\d+)', line).group(1)
            indexone = map_addr_to_index(int(value_after_addr))
            if str(indextwo) in parsed_data[0]:
                data_templates["node_script_sending"][indexone][parsed_data[0].index(str(indextwo))]=time
# 处理包含 "===============on bm，" 的行
def process_on_bm(line):
    parts = line.split('msg: ', 1)
    if len(parts) > 1:
        msg = parts[1]
        if "ext" in msg:
            json_str = re.search(r'({.*})', msg).group(1)
            on_bm = json.loads(json_str)
            value_after_ext = re.search(r'"ext"\s*:"\s*(\d+)"', msg).group(1)
            indexadd = ext_mapping[value_after_ext]
            indextwo = addr_mapping[int(indexadd)]
            value_mapping[indexadd] = on_bm["value"][0]
            value_last_mapping[str(int(indexadd)+len(on_bm["value"])-1)] = on_bm["value"][len( on_bm["value"])-1]
            time = re.search(r'time:(\d+)', line).group(1)
            indexone = map_addr_to_index(int(indexadd))
            if str(indextwo) in parsed_data[0]:
                data_templates["node_script_receiving"][indexone][parsed_data[0].index(str(indextwo))]=time
# 处理包含 "=============== 脚本" 和 "收到" 的行
def process_script_receive(line):
    json_str = re.search(r'\[({.*?})\]', line).group(0)
    onsend = json.loads(json_str)[0]["fullName"]
    value_after_addr = re.search(r'[YX](\d+)', onsend).group(1)
    if int(value_after_addr) in addr_mapping:
        indextwo = addr_mapping[int(value_after_addr)]
        time = re.search(r'time:(\d+)', line).group(1)
        if str(indextwo) in parsed_data[0]:
            if "y1" in line:
                data_templates["front_end_script_sending_0x0d"][0][parsed_data[0].index(str(indextwo))]=time            
            elif "y2" in line:
                data_templates["front_end_script_sending_0x0d"][1][parsed_data[0].index(str(indextwo))]=time            
            elif "y3" in line:
                data_templates["front_end_script_sending_0x0d"][2][parsed_data[0].index(str(indextwo))]=time            
            elif "y4" in line:
                data_templates["front_end_script_sending_0x0d"][3][parsed_data[0].index(str(indextwo))]=time            
            elif "x" in line:
                data_templates["front_end_script_sending_0x0d"][4][parsed_data[0].index(str(indextwo))]=time
# 处理包含 "=============== send BatchSetByFullNames，" 的行
def process_send_batch_set_by_full_names(line):
    parts = line.split('msg: ', 1)
    if len(parts) > 1:
        msg = parts[1]
        if "data" in msg:
            json_str = re.search(r'({.*})', msg).group(1)
            onsend = json.loads(json_str)["data"]
            id = onsend.split("|@", 1)[0]
            if int(id) in id_mapping:
                indextwo = id_mapping[int(id)]
                time = re.search(r'time:(\d+)', line).group(1)
                indexone = map_id_to_index(int(id))
                if str(indextwo) in parsed_data[0]:
                    data_templates["node_sending_0x0d"][indexone][parsed_data[0].index(str(indextwo))]=time
# 处理包含 "=============== tcp receiveData，" 的行
def process_tcp_receive_data(line):
    parts = line.split('msg: ', 1)
    if len(parts) > 1:
        msg = parts[1]
        if "16059" in msg:
            value_after_16059 = (re.search(r'16059\|@(\d+)', msg).group(1))
            parsed_data[0][int(value_after_16059)+1]=(value_after_16059)
            time = re.search(r'time:(\d+)', line).group(1)
            parsed_data[3][parsed_data[0].index((value_after_16059))]=(time)
        process_ext_fields(msg, line)
        process_id_fields(msg, line)

# 处理 ext 字段
def process_ext_fields(msg, line):
    match_ext = re.findall(r'"ext":"(\d+)"', msg)
    if match_ext:
        for ext in match_ext:
            indexadd = ext_mapping[ext]
            indextwo = addr_mapping[int(indexadd)]
            time = re.search(r'time:(\d+)', line).group(1)
            indexone = map_addr_to_index(int(indexadd))
            if str(indextwo) in parsed_data[0]:
                data_templates["www_receive_data"][indexone][parsed_data[0].index(str(indextwo))]=time
# 处理 id 字段
def process_id_fields(msg, line):
    for key in id_mapping:
        plc=find_key_by_value(plc_mapping,key)
        if len(value_mapping)>0 and str(plc) in value_mapping:
            value=value_mapping[str(plc)]
            if (str(key) + "|@"+str(value)) in msg:
                indextwo = id_mapping[int(key)]
                time = re.search(r'time:(\d+)', line).group(1)
                indexone = map_id_to_index(key)
                parsed_data_index = parsed_data[0].index(str(indextwo))
                data_templates["www_receive_0x01"][indexone][parsed_data_index]=time

# 处理包含 "=============== useVarTrigger 35537 ，" 的行
def process_use_var_trigger(line):
    parts = re.search(r'msg:\s*(\d+)', line).group(1)
    if len(parts) > 0:
        time = re.search(r'time:(\d+)', line).group(1)
        if str(parts) in parsed_data[0]:
            data_templates["front_end_receive_35537"][0][parsed_data[0].index(str(parts))]=time

def process_render(line):
    parts = re.search(r'msg:\s*(\d+)', line).group(1)
    if len(parts) > 0:
        time = re.search(r'time:(\d+)', line).group(1)
        if str(parts) in parsed_data[0]:
            data_templates["front_end_receive_35537"][1][parsed_data[0].index(str(parts))]=time

def merge_data_templates():
    for key in data_templates:
        parsed_data.extend(data_templates[key])

def save_to_excel():
    for index, line in enumerate(parsed_data):
        sheet.append(line)
        if any("采集" in str(cell) for cell in line):
            for cell in sheet[index + 1]:
                cell.fill = highlight_fill
    output_filename = LOG_FILE_PATH.split("/")[-1].replace(".log", "_parsed.xlsx")
    workbook.save("parsed.xlsx")
    print(f"数据已成功提取并保存到 Excel 文件: {output_filename}")

process_log_file()